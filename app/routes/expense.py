"""Expense workflow routes (电费支出给国网) - PostgreSQL"""
import os, re, shutil, zipfile, tempfile, hashlib
from datetime import datetime
from flask import Blueprint, request, jsonify
import openpyxl
from config import Config, BASE_DIR
from app.database import db_session
from app.routes.auth_helper import user_id, user_name
from app.utils.helpers import sanitize_filename, decode_filename
from app.services.email_client import EmailClient
from app.services.file_manager import FileManager
from app.services.excel_handler import ExcelHandler

expense_bp = Blueprint('expense', __name__)


def _json(data, code=200):
    return jsonify(data), code


def _workspace_dir():
    """Get the current user's workspace directory."""
    uname = user_name()
    if uname:
        d = os.path.join(Config.ATTACHMENTS_DIR, uname)
        os.makedirs(d, exist_ok=True)
        return d
    return Config.ATTACHMENTS_DIR


# Email endpoints (no user_id needed)
@expense_bp.route('/api/expense_scan', methods=['POST'])
def expense_scan():
    try:
        data = request.get_json(force=True)
        client = EmailClient(address=data.get('email') or None, password=data.get('password') or None,
                             server=data.get('server') or None, port=int(data.get('port', 993)))
        success, logs, total, bill_count, invoice_count = client.scan_emails(
            data.get('start_date', ''), data.get('end_date', ''))
        return _json({'success': success, 'logs': logs, 'total': total, 'bill_count': bill_count, 'invoice_count': invoice_count}, 500 if not success else 200)
    except Exception as e:
        import traceback
        return _json({'success': False, 'logs': [f'扫描出错: {e}', traceback.format_exc()[-200:]]}, 500)


@expense_bp.route('/api/expense_download_bills', methods=['POST'])
def expense_download_bills():
    try:
        data = request.get_json(force=True)
        output_dir = data.get('output_dir', _workspace_dir())
        client = EmailClient(address=data.get('email') or None, password=data.get('password') or None,
                             server=data.get('server') or None, port=int(data.get('port', 993)))
        success, logs, count = client.download_bills(data.get('start_date', ''), data.get('end_date', ''), output_dir)
        pdfs = sorted([f for f in os.listdir(output_dir) if f.endswith('.pdf') and '用户[' in f]) if os.path.isdir(output_dir) else []
        return _json({'success': success, 'logs': logs, 'count': count, 'html_url': '/attachments/bill_download.html' if pdfs else ''}, 500 if not success else 200)
    except Exception as e:
        import traceback
        return _json({'success': False, 'logs': [f'下载出错: {e}', traceback.format_exc()[-300:]]}, 500)


@expense_bp.route('/api/expense_download_invoices', methods=['POST'])
def expense_download_invoices():
    data = request.get_json(force=True)
    client = EmailClient(address=data.get('email') or None, password=data.get('password') or None,
                         server=data.get('server') or None, port=int(data.get('port', 993)))
    success, logs, links = client.collect_invoice_links(data.get('start_date', ''), data.get('end_date', ''))
    if links:
        output_dir = data.get('output_dir', _workspace_dir())
        os.makedirs(output_dir, exist_ok=True)
        html = '<!DOCTYPE html><html><head><meta charset="UTF-8"><title>发票下载</title></head><body>'
        html += f'<h2>{len(links)} 个发票链接</h2>'
        for i, link in enumerate(links):
            html += f'<p><a href="{link}" target="_blank">#{i+1} 下载发票</a></p>'
        html += '</body></html>'
        with open(os.path.join(output_dir, 'invoice_download.html'), 'w', encoding='utf-8') as f:
            f.write(html)
    return _json({'success': success, 'logs': logs, 'count': len(links), 'html_url': '/attachments/invoice_download.html' if links else ''}, 500 if not success else 200)


@expense_bp.route('/api/expense_copy_invoices', methods=['POST'])
def expense_copy_invoices():
    from app.services.pdf_extractor import PDFExtractor
    data = request.get_json(force=True)
    output_dir = data.get('output_dir', _workspace_dir())
    dl_dir = data.get('downloads_dir') or Config.DOWNLOADS_DIR
    logs, count = FileManager.copy_invoices_from_downloads(output_dir, dl_dir)
    success = count > 0
    if success:
        r_logs, _ = FileManager.rename_all_invoices(output_dir)
        logs.extend(r_logs)
    return _json({'success': success, 'logs': logs, 'count': count})


@expense_bp.route('/api/expense_rename_invoices', methods=['POST'])
def expense_rename_invoices():
    data = request.get_json(force=True)
    input_dir = data.get('output_dir', _workspace_dir())
    logs, count = FileManager.rename_all_invoices(input_dir)
    success = count > 0 or len(logs) == 0
    return _json({'success': success, 'logs': logs, 'total': 0, 'count': count})


@expense_bp.route('/api/expense_extract', methods=['POST'])
def expense_extract():
    from app.services.pdf_extractor import PDFExtractor
    data = request.get_json(force=True)
    output_dir = data.get('output_dir', _workspace_dir())
    mapping_path = data.get('mapping_path', Config.MAPPING_FILE)
    logs = []
    if not os.path.exists(output_dir):
        logs.append(f'错误: 目录不存在: {output_dir}')
        return _json({'success': False, 'logs': logs, 'count': 0})
        return _json({'success': False, 'logs': logs, 'count': 0})

    shortname_mapping = FileManager.load_shortname_mapping(mapping_path)
    logs.append(f'加载 {len(shortname_mapping)} 条简称映射')

    pdf_files = [fn for fn in os.listdir(output_dir) if fn.lower().endswith('.pdf') and '用户[' in fn]
    if not pdf_files:
        logs.append('错误: 未找到电费单PDF文件')
        return _json({'success': False, 'logs': logs, 'count': 0})

    # Limit to avoid Render timeout (200 max)
    if len(pdf_files) > 200:
        logs.append(f'PDF较多({len(pdf_files)}个)，分批处理前200个')
        pdf_files = pdf_files[:200]

    logs.append(f'找到 {len(pdf_files)} 个电费单PDF文件')
    data_list = []
    for idx, fn in enumerate(pdf_files, 1):
        logs.append(f'处理 {idx}/{len(pdf_files)}: {fn}')
        try:
            info = PDFExtractor.extract_bill_data(os.path.join(output_dir, fn))
        except Exception as e:
            logs.append(f'  提取异常: {e}')
            continue
        if not info:
            logs.append('  提取失败')
            continue
        uid = info.get('user_id', '')
        user_key = f'[{uid}]'
        shortname = shortname_mapping.get(user_key, '未知')
        sn_match = re.search(r'\((.*?)\)\.pdf', fn)
        if sn_match:
            shortname = sn_match.group(1)
        data_list.append({'用户号': uid, '简称': shortname, '电费年月': info.get('year_month', ''),
                          '本期电费(元)': info.get('electricity_fee', ''), '用电量(度)': info.get('electricity_kwh', ''),
                          '文件名': fn})

    if not data_list:
        logs.append('错误: 未提取到任何数据')
        return _json({'success': False, 'logs': logs, 'count': 0})

    output_file = ExcelHandler.generate_summary(data_list, os.path.join(Config.ATTACHMENTS_DIR, f'电费单信息汇总-{datetime.now().strftime("%Y年%m月%d日")}.xlsx'))
    logs.append(f'共提取 {len(data_list)} 条记录')
    return _json({'success': True, 'logs': logs, 'count': len(data_list), 'file': output_file})


@expense_bp.route('/api/expense_add_shortname', methods=['POST'])
def expense_add_shortname():
    data = request.get_json(force=True)
    input_dir = data.get('input_dir', _workspace_dir())
    mapping_path = data.get('mapping_path', Config.MAPPING_FILE)
    logs, count = FileManager.add_shortname_to_files(input_dir, mapping_path)
    return _json({'success': count > 0, 'logs': logs, 'count': count})


@expense_bp.route('/api/expense_classify', methods=['POST'])
def expense_classify():
    data = request.get_json(force=True)
    input_dir = data.get('input_dir', _workspace_dir())
    mapping_path = data.get('mapping_path', Config.MAPPING_FILE)
    logs, count = FileManager.classify_files(input_dir, mapping_path)
    return _json({'success': count > 0, 'logs': logs, 'count': count})


@expense_bp.route('/api/expense_dedup', methods=['POST'])
def expense_dedup():
    data = request.get_json(force=True)
    input_dir = data.get('input_dir', _workspace_dir())
    logs = []
    if not os.path.exists(input_dir):
        return _json({'success': False, 'logs': ['目录不存在'], 'count': 0})
    seen = {}
    kept = removed = 0
    for fn in sorted(os.listdir(input_dir)):
        if not fn.lower().endswith('.pdf'):
            continue
        fp_full = os.path.join(input_dir, fn)
        with open(fp_full, 'rb') as f:
            h = hashlib.md5(f.read()).hexdigest()
        if h in seen:
            os.remove(fp_full)
            logs.append(f'× 重复删除: {fn}')
            removed += 1
        else:
            seen[h] = fn; kept += 1
    logs.append(f'去重完成: 保留 {kept} 个, 删除 {removed} 个')
    return _json({'success': True, 'logs': logs, 'count': removed, 'kept': kept})


@expense_bp.route('/api/expense_upload', methods=['POST'])
def expense_upload_expense():
    import glob as gb
    # Search in attachments dir first (workspace), then root
    files = gb.glob(os.path.join(_workspace_dir(), '电费单信息汇总-*.xlsx'))
    if not files:
        files = gb.glob(os.path.join(Config.ATTACHMENTS_DIR, '电费单信息汇总-*.xlsx'))
    if not files:
        return _json({'success': False, 'logs': ['未找到汇总文件'], 'count': 0})
    xlsx_file = sorted(files)[-1]
    uid = user_id()
    logs = [f'读取: {os.path.basename(xlsx_file)}']
    try:
        wb = openpyxl.load_workbook(xlsx_file, data_only=True)
        ws = wb.active
        headers = [str(c.value) if c.value else '' for c in ws[1]]
        uid_col = month_col = fee_col = kwh_col = None
        for i, h in enumerate(headers):
            if '用户号' in h or '户号' in h: uid_col = i
            elif '周期' in h or '年月' in h: month_col = i
            elif '电费' in h: fee_col = i
            elif '电量' in h or '用电' in h: kwh_col = i
        if uid_col is None or month_col is None:
            return _json({'success': False, 'logs': ['未找到用户号或年月列'], 'count': 0})
        with db_session() as conn:
            c = conn.cursor()
            new_count = dup_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                meter_raw = str(row[uid_col] or '').strip()
                if not meter_raw or meter_raw == 'None': continue
                meter_id = meter_raw.replace('[', '').replace(']', '')
                month_raw = str(row[month_col] or '').strip()
                m = re.search(r'(\d{4})\s*年\s*(\d{1,2})', month_raw)
                if not m: continue
                year_month = f'{m.group(1)}-{int(m.group(2)):02d}'
                fee = kwh = 0.0
                if fee_col is not None and row[fee_col]:
                    try: fee = float(str(row[fee_col]).replace(',', ''))
                    except: pass
                if kwh_col is not None and row[kwh_col]:
                    try: kwh = float(str(row[kwh_col]).replace(',', ''))
                    except: pass
                if fee == 0 and kwh == 0: continue
                c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=%s AND year_month=%s AND category='电费支' AND user_id=%s", (meter_id, year_month, uid))
                if c.fetchone()[0] == 0 and fee > 0:
                    c.execute("INSERT INTO transactions (meter_id, year_month, category, amount, user_id) VALUES (%s, %s, '电费支', %s, %s)", (meter_id, year_month, fee, uid))
                    new_count += 1
                else: dup_count += 1
                c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=%s AND year_month=%s AND category='电量支' AND user_id=%s", (meter_id, year_month, uid))
                if c.fetchone()[0] == 0 and kwh > 0:
                    c.execute("INSERT INTO transactions (meter_id, year_month, category, amount, user_id) VALUES (%s, %s, '电量支', %s, %s)", (meter_id, year_month, kwh, uid))
                    new_count += 1
                else: dup_count += 1
            conn.commit()
        logs.append(f'上传完成！新增 {new_count} 条，跳过 {dup_count} 条')
        return _json({'success': True, 'logs': logs, 'new': new_count, 'dup': dup_count})
    except Exception as e:
        logs.append(f'上传失败: {e}')
        return _json({'success': False, 'logs': logs, 'count': 0}, 500)


@expense_bp.route('/api/expense_upload_file', methods=['POST'])
def expense_upload_file():
    import base64
    data = request.get_json(force=True)
    uid = user_id()
    logs = []
    b64 = data.get('data', '')
    if not b64:
        return _json({'success': False, 'logs': ['没有文件数据']})
    raw = base64.b64decode(b64)
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    tmp.write(raw); tmp.close()
    logs.append(f'接收文件: {data.get("name", "")}')
    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        ws = wb.active
        headers = [str(c.value) if c.value else '' for c in ws[1]]
        uid_col = month_col = fee_col = kwh_col = None
        for i, h in enumerate(headers):
            if '用户号' in h or '户号' in h: uid_col = i
            elif '周期' in h or '年月' in h: month_col = i
            elif '电费' in h: fee_col = i
            elif '电量' in h: kwh_col = i
        if uid_col is None or month_col is None:
            os.unlink(tmp.name)
            return _json({'success': False, 'logs': ['未找到用户号或年月列'], 'count': 0})
        with db_session() as conn:
            c = conn.cursor()
            new_count = dup_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                meter_raw = str(row[uid_col] or '').strip()
                if not meter_raw or meter_raw == 'None': continue
                meter_id = meter_raw.replace('[', '').replace(']', '')
                month_raw = str(row[month_col] or '').strip()
                m = re.search(r'(\d{4})\s*年\s*(\d{1,2})', month_raw)
                if not m: continue
                year_month = f'{m.group(1)}-{int(m.group(2)):02d}'
                fee = kwh = 0.0
                if fee_col and row[fee_col]: 
                    try: fee = float(str(row[fee_col]).replace(',', ''))
                    except: pass
                if kwh_col and row[kwh_col]:
                    try: kwh = float(str(row[kwh_col]).replace(',', ''))
                    except: pass
                if fee == 0 and kwh == 0: continue
                c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=%s AND year_month=%s AND category='电费支' AND user_id=%s", (meter_id, year_month, uid))
                if c.fetchone()[0] == 0 and fee > 0:
                    c.execute("INSERT INTO transactions (meter_id, year_month, category, amount, user_id) VALUES (%s, %s, '电费支', %s, %s)", (meter_id, year_month, fee, uid))
                    new_count += 1
                else: dup_count += 1
                c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=%s AND year_month=%s AND category='电量支' AND user_id=%s", (meter_id, year_month, uid))
                if c.fetchone()[0] == 0 and kwh > 0:
                    c.execute("INSERT INTO transactions (meter_id, year_month, category, amount, user_id) VALUES (%s, %s, '电量支', %s, %s)", (meter_id, year_month, kwh, uid))
                    new_count += 1
                else: dup_count += 1
            conn.commit()
        os.unlink(tmp.name)
        logs.append(f'上传完成！新增 {new_count} 条，跳过 {dup_count} 条')
        return _json({'success': True, 'logs': logs, 'new': new_count, 'dup': dup_count})
    except Exception as e:
        os.unlink(tmp.name)
        logs.append(f'上传失败: {e}')
        return _json({'success': False, 'logs': logs}, 500)


@expense_bp.route('/api/expense_upload_invoices', methods=['POST'])
def expense_upload_invoices():
    from app.services.pdf_extractor import PDFExtractor
    import base64
    data = request.get_json(force=True)
    logs = []
    files = data.get('files', [])
    if not files:
        return _json({'success': False, 'logs': ['没有接收到文件']})
    save_dir = data.get('output_dir', _workspace_dir())
    os.makedirs(save_dir, exist_ok=True)
    saved = 0
    for f_info in files:
        fn = os.path.basename(f_info.get('name', 'invoice.pdf'))
        b64 = f_info.get('data', '')
        if not b64: continue
        payload = base64.b64decode(b64)
        save_path = os.path.join(save_dir, fn)
        c = 1
        while os.path.exists(save_path):
            b, e2 = os.path.splitext(fn)
            save_path = os.path.join(save_dir, f'{b}_{c}{e2}'); c += 1
        with open(save_path, 'wb') as f:
            f.write(payload)
        saved += 1
        logs.append(f'接收: {fn}')
    if saved > 0:
        logs.append('开始重命名...')
        renamed = 0
        for fn in sorted(os.listdir(save_dir)):
            if not fn.endswith('.pdf'): continue
            if not fn.startswith('sdp_'): continue
            success, result = PDFExtractor.rename_invoice(os.path.join(save_dir, fn))
            if success:
                logs.append(f'重命名: {fn} -> {result}')
                renamed += 1
        logs.append(f'重命名完成: {renamed} 个')
    return _json({'success': True, 'logs': logs, 'count': saved})


@expense_bp.route('/api/expense_list_files')
def expense_list_files():
    dir_path = _workspace_dir()
    files = []
    if os.path.isdir(dir_path):
        for root, dirs, fns in os.walk(dir_path):
            for fn in fns:
                if fn.startswith('_'): continue
                fp_full = os.path.join(root, fn)
                rel = os.path.relpath(fp_full, dir_path).replace('\\', '/')
                files.append({'name': rel, 'size': os.path.getsize(fp_full)})
    return jsonify({'files': files})


@expense_bp.route('/api/expense_pack')
def expense_pack():
    dir_path = _workspace_dir()
    if not os.path.isdir(dir_path): return '', 404
    zip_path = os.path.join(dir_path, '_pack.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for root, dirs, files in os.walk(dir_path):
            for fn in files:
                if fn == '_pack.zip': continue
                zf.write(os.path.join(root, fn), os.path.relpath(os.path.join(root, fn), dir_path))
    return jsonify({'url': '/attachments/_pack.zip'})


@expense_bp.route('/api/expense_pack_selected', methods=['POST'])
def expense_pack_selected():
    data = request.get_json(force=True)
    names = data.get('files', [])
    dir_path = _workspace_dir()
    zip_path = os.path.join(dir_path, '_selected.zip')
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fn in names:
            fp_full = os.path.join(dir_path, fn)
            if os.path.isfile(fp_full): zf.write(fp_full, fn)
    return jsonify({'success': True})


@expense_bp.route('/api/expense_cleanup_temp', methods=['POST'])
def expense_cleanup_temp():
    for fn in ['_browse.html', '_selected.zip']:
        p = os.path.join(_workspace_dir(), fn)
        if os.path.isfile(p):
            try: os.unlink(p)
            except: pass
    return jsonify({'success': True})


@expense_bp.route('/api/expense_clear', methods=['POST'])
def expense_clear():
    data = request.get_json(force=True)
    output_dir = data.get('output_dir', _workspace_dir())
    logs, count = FileManager.clear_attachments(output_dir)
    try:
        src = os.path.join(Config.STATIC_DIR, '_browse.html')
        dst = os.path.join(output_dir, '_browse.html')
        if os.path.isfile(src): shutil.copy2(src, dst)
    except: pass
    return _json({'success': True, 'logs': logs, 'count': count})
