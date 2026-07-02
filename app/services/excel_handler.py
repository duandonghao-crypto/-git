"""
Excel handler — read/write summary spreadsheets and upload data to DB.
"""
import os
import re
import glob
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill
from config import Config
from app.database import db_session


class ExcelHandler:
    """Handle electricity bill data Excel operations."""

    @staticmethod
    def generate_summary(data_list: list, output_file: str = None) -> str:
        """Generate an Excel summary file from extracted bill data.
        Returns the output file path."""
        today = datetime.now().strftime('%Y年%m月%d日')
        if not output_file:
            output_file = os.path.join(Config.ATTACHMENTS_DIR,
                                       f'电费单信息汇总-{today}.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '电费单信息'
        headers = ['用户号', '简称', '电费年月', '本期电费(元)', '用电量(度)', '文件名']

        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.font = header_font
            cell.fill = header_fill

        for ri, d in enumerate(data_list, 2):
            ws.cell(row=ri, column=1, value=d.get('用户号') or d.get('user_id', ''))
            ws.cell(row=ri, column=2, value=d.get('简称', ''))
            ws.cell(row=ri, column=3, value=d.get('电费年月') or d.get('year_month', ''))
            ws.cell(row=ri, column=4, value=d.get('本期电费(元)') or d.get('electricity_fee', ''))
            ws.cell(row=ri, column=5, value=d.get('用电量(度)') or d.get('electricity_kwh', ''))
            ws.cell(row=ri, column=6, value=d.get('文件名') or d.get('filename', ''))

        for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F'], [15, 25, 15, 18, 15, 45]):
            ws.column_dimensions[col].width = width

        wb.save(output_file)
        return output_file

    @staticmethod
    def upload_expense_from_excel(xlsx_file: str):
        """Upload expense data from the summary Excel to DB as 电费支/电量支."""
        logs = []
        if not os.path.exists(xlsx_file):
            logs.append(f"文件不存在: {xlsx_file}")
            return False, logs, 0, 0

        logs.append(f"读取文件: {os.path.basename(xlsx_file)}")
        try:
            wb = openpyxl.load_workbook(xlsx_file, data_only=True)
            ws = wb.active
            headers = [str(c.value) if c.value else '' for c in ws[1]]
            logs.append(f"表头: {headers}")

            uid_col = month_col = fee_col = kwh_col = None
            for i, h in enumerate(headers):
                if '用户号' in h or '户号' in h:
                    uid_col = i
                elif '周期' in h or '年月' in h:
                    month_col = i
                elif '电费' in h:
                    fee_col = i
                elif '电量' in h or '用电' in h:
                    kwh_col = i

            if uid_col is None or month_col is None:
                logs.append("错误: 未找到用户号或年月列")
                return False, logs, 0, 0

            with db_session() as conn:
                c = conn.cursor()
                new_count = 0
                dup_count = 0

                for row in ws.iter_rows(min_row=2, values_only=True):
                    meter_raw = str(row[uid_col] or '').strip()
                    if not meter_raw or meter_raw == 'None':
                        continue
                    meter_id = meter_raw.replace('[', '').replace(']', '')

                    month_raw = str(row[month_col] or '').strip()
                    m_m = re.search(r'(\d{4})\s*年\s*(\d{1,2})', month_raw)
                    if not m_m:
                        continue
                    year_month = f'{m_m.group(1)}-{int(m_m.group(2)):02d}'

                    fee = 0.0
                    if fee_col is not None and row[fee_col] is not None:
                        try:
                            fee = float(str(row[fee_col]).replace(',', ''))
                        except (ValueError, TypeError):
                            pass

                    kwh = 0.0
                    if kwh_col is not None and row[kwh_col] is not None:
                        try:
                            kwh = float(str(row[kwh_col]).replace(',', ''))
                        except (ValueError, TypeError):
                            pass

                    if fee == 0 and kwh == 0:
                        continue

                    c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=? AND year_month=? AND category='电费支'",
                              (meter_id, year_month))
                    fee_exists = c.fetchone()[0] > 0
                    c.execute("SELECT COUNT(*) FROM transactions WHERE meter_id=? AND year_month=? AND category='电量支'",
                              (meter_id, year_month))
                    kwh_exists = c.fetchone()[0] > 0

                    if fee > 0:
                        if fee_exists:
                            dup_count += 1
                        else:
                            c.execute("INSERT INTO transactions (meter_id, year_month, category, amount) VALUES (?, ?, '电费支', ?)",
                                      (meter_id, year_month, fee))
                            new_count += 1
                    if kwh > 0:
                        if kwh_exists:
                            dup_count += 1
                        else:
                            c.execute("INSERT INTO transactions (meter_id, year_month, category, amount) VALUES (?, ?, '电量支', ?)",
                                      (meter_id, year_month, kwh))
                            new_count += 1

                conn.commit()
            logs.append(f"上传完成！新增 {new_count} 条，跳过 {dup_count} 条重复")
            return True, logs, new_count, dup_count
        except Exception as e:
            logs.append(f"上传失败: {str(e)}")
            return False, logs, 0, 0
