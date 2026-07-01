"""
File manager — classify, dedup, rename files in attachments directory.
"""
import os
import re
import shutil
import hashlib
import subprocess
import openpyxl
from config import Config


class FileManager:
    """Manage files in the attachments directory."""

    @staticmethod
    def load_shortname_mapping(mapping_path=None):
        """Load user ID -> short name from Excel. Column A=user ID, Column B=short name."""
        mapping_path = mapping_path or Config.MAPPING_FILE
        abs_path = os.path.abspath(mapping_path)
        if not os.path.exists(abs_path):
            return {}
        try:
            mapping = {}
            wb = openpyxl.load_workbook(abs_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 2 and row[0] is not None and row[1] is not None:
                    uid = str(row[0]).strip()
                    sname = str(row[1]).strip()
                    if uid and sname and uid not in ('nan', 'None') and sname not in ('nan', 'None'):
                        if not uid.startswith('['):
                            uid = f'[{uid}]'
                        mapping[uid] = sname
            wb.close()
            return mapping
        except Exception:
            return {}

    @staticmethod
    def load_company_mapping(mapping_path=None):
        """Load user ID -> company name mapping from Excel.
        Column A: user ID (户号/用户号), Column C: company name (公司)."""
        mapping_path = mapping_path or Config.COMPANY_MAPPING_FILE
        if not os.path.exists(mapping_path):
            return {}
        try:
            mapping = {}
            wb = openpyxl.load_workbook(mapping_path, data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                if len(row) >= 3 and row[0] is not None and row[2] is not None:
                    uid = str(row[0]).strip().replace(' ', '').replace('\t', '')
                    company = str(row[2]).strip()
                    if uid and company and uid != 'None' and company != 'None':
                        if not uid.startswith('['):
                            uid = f'[{uid}]'
                        mapping[uid] = company
            wb.close()
            return mapping
        except Exception:
            return {}

    @staticmethod
    def add_shortname_to_files(input_dir=None, mapping_path=None):
        """Rename PDF files by appending short name from mapping."""
        input_dir = input_dir or Config.ATTACHMENTS_DIR
        mapping_path = mapping_path or Config.MAPPING_FILE
        logs = []

        if not os.path.exists(input_dir):
            logs.append(f"错误: 目录不存在: {input_dir}")
            return logs, 0

        if not os.path.exists(mapping_path):
            logs.append(f"错误: 映射文件不存在: {mapping_path}")
            return logs, 0

        try:
            mapping = FileManager.load_shortname_mapping(mapping_path)
            logs.append(f"已加载 {len(mapping)} 条映射")
            count = 0

            for fn in sorted(os.listdir(input_dir)):
                if not fn.lower().endswith('.pdf'):
                    continue
                m = re.search(r'\[([^\]]+)\]', fn)
                if not m:
                    logs.append(f"跳过（无户号）: {fn}")
                    continue
                key = m.group(0)
                short = mapping.get(key, '')
                if not short:
                    logs.append(f"跳过（无简称）: {fn}")
                    continue

                old_path = os.path.join(input_dir, fn)
                parts = fn.rsplit('.', 1)
                new_fn = f"{parts[0]}({short}).{parts[1]}"
                new_path = os.path.join(input_dir, new_fn)
                os.rename(old_path, new_path)
                logs.append(f"✓ {fn} → {new_fn}")
                count += 1

            return logs, count
        except Exception as e:
            logs.append(f"错误: {str(e)}")
            return logs, 0

    @staticmethod
    def classify_files(input_dir=None, mapping_path=None):
        """Classify bills and invoices into company subdirectories."""
        input_dir = input_dir or Config.ATTACHMENTS_DIR
        mapping_path = mapping_path or Config.MAPPING_FILE
        logs = []

        if not os.path.exists(input_dir):
            logs.append(f"错误: 目录不存在: {input_dir}")
            return logs, 0

        if not os.path.exists(mapping_path):
            logs.append(f"错误: 映射文件不存在: {mapping_path}")
            return logs, 0

        try:
            company_map = FileManager.load_company_mapping(mapping_path)
            logs.append(f"已加载 {len(company_map)} 条户号→公司映射")

            # Classify bills
            bills = [f for f in os.listdir(input_dir)
                     if f.lower().endswith('.pdf') and '用户[' in f]
            logs.append(f"找到 {len(bills)} 个电费账单")
            bill_count = 0

            for fn in bills:
                m = re.search(r'用户\[(\d+)\]', fn)
                if not m:
                    continue
                uid = f'[{m.group(1)}]'
                company = company_map.get(uid)
                if not company:
                    logs.append(f"跳过 {fn}: 未找到公司映射")
                    continue
                company_dir = os.path.join(input_dir, company)
                os.makedirs(company_dir, exist_ok=True)
                old_p = os.path.join(input_dir, fn)
                new_p = os.path.join(company_dir, fn)
                if not os.path.exists(new_p):
                    os.rename(old_p, new_p)
                    bill_count += 1
                    logs.append(f"✓ 账单: {fn} → {company}/")

            # Classify invoices
            invoices = [f for f in os.listdir(input_dir)
                        if f.lower().endswith('.pdf') and '电费发票' in f]
            logs.append(f"找到 {len(invoices)} 个发票")
            inv_count = 0

            for fn in invoices:
                m = re.match(r'^(.+?有限公司)(\d+)月电费发票.*\.pdf$', fn)
                if not m:
                    continue
                company = m.group(1).strip()
                company_dir = os.path.join(input_dir, company)
                os.makedirs(company_dir, exist_ok=True)
                old_p = os.path.join(input_dir, fn)
                new_p = os.path.join(company_dir, fn)
                if not os.path.exists(new_p):
                    os.rename(old_p, new_p)
                    inv_count += 1
                    logs.append(f"✓ 发票: {fn} → {company}/")

            total = bill_count + inv_count
            logs.append(f"分类完成: 账单 {bill_count} 个 + 发票 {inv_count} 个 = 共 {total} 个")
            return logs, total
        except Exception as e:
            logs.append(f"错误: {str(e)}")
            return logs, 0

    @staticmethod
    def dedup_files(input_dir=None):
        """Remove duplicate PDFs by MD5 hash."""
        input_dir = input_dir or Config.ATTACHMENTS_DIR
        logs = []
        if not os.path.exists(input_dir):
            logs.append(f"错误: 目录不存在: {input_dir}")
            return logs, 0, 0

        seen = {}
        kept = 0
        removed = 0
        for fn in sorted(os.listdir(input_dir)):
            if not fn.lower().endswith('.pdf'):
                continue
            fp_full = os.path.join(input_dir, fn)
            with open(fp_full, 'rb') as f:
                h = hashlib.md5(f.read()).hexdigest()
            if h in seen:
                os.remove(fp_full)
                logs.append(f"× 重复删除: {fn} (同 {seen[h]})")
                removed += 1
            else:
                seen[h] = fn
                kept += 1
        logs.append(f"去重完成: 保留 {kept} 个, 删除 {removed} 个重复")
        return logs, kept, removed

    @staticmethod
    def clear_attachments(output_dir=None):
        """Clear all files in the attachments directory."""
        output_dir = output_dir or Config.ATTACHMENTS_DIR
        logs = []
        if not os.path.exists(output_dir):
            logs.append("目录不存在，无需清空")
            return logs, 0

        count = 0
        for item in os.listdir(output_dir):
            item_path = os.path.join(output_dir, item)
            try:
                if os.path.isfile(item_path):
                    try:
                        os.unlink(item_path)
                    except Exception:
                        subprocess.run(['cmd', '/c', 'del', '/f', '/q', item_path],
                                       capture_output=True, timeout=10)
                    count += 1
                elif os.path.isdir(item_path):
                    subprocess.run(['cmd', '/c', 'rmdir', '/s', '/q', item_path],
                                   capture_output=True, timeout=30)
                    count += 1
            except Exception as e:
                logs.append(f"失败 {item}: {e}")
        logs.append(f"删除 {count} 个项目")
        return logs, count

    @staticmethod
    def copy_invoices_from_downloads(target_dir=None, downloads_dir=None):
        """Copy downloaded sdp_*.pdf files from Downloads to attachments."""
        target_dir = target_dir or Config.ATTACHMENTS_DIR
        downloads_dir = downloads_dir or Config.DOWNLOADS_DIR
        logs = []

        if not os.path.exists(downloads_dir):
            logs.append(f"错误: Downloads目录不存在: {downloads_dir}")
            return logs, 0

        os.makedirs(target_dir, exist_ok=True)
        sdp_files = [f for f in os.listdir(downloads_dir)
                     if f.startswith('sdp_') and f.lower().endswith('.pdf')]

        if not sdp_files:
            logs.append("未找到sdp_开头的发票PDF文件")
            return logs, 0

        logs.append(f"找到 {len(sdp_files)} 个发票PDF")
        count = 0
        for fn in sdp_files:
            src = os.path.join(downloads_dir, fn)
            dst = os.path.join(target_dir, fn)
            if os.path.exists(dst):
                base, ext = os.path.splitext(fn)
                c = 1
                while os.path.exists(os.path.join(target_dir, f"{base}_{c}{ext}")):
                    c += 1
                dst = os.path.join(target_dir, f"{base}_{c}{ext}")
            try:
                shutil.copy2(src, dst)
                count += 1
                logs.append(f"✓ 导入: {fn}")
            except Exception as e:
                logs.append(f"导入失败 {fn}: {e}")

        logs.append(f"导入完成！共导入 {count} 个发票")
        return logs, count

    @staticmethod
    def rename_all_invoices(input_dir=None):
        """Rename all sdp_*.pdf files in directory to 公司名X月电费发票.pdf."""
        from app.services.pdf_extractor import PDFExtractor
        input_dir = input_dir or Config.ATTACHMENTS_DIR
        logs = []
        if not os.path.exists(input_dir):
            logs.append(f"错误: 目录不存在: {input_dir}")
            return logs, 0

        renamed = 0
        for fn in sorted(os.listdir(input_dir)):
            if not (fn.startswith('sdp_') and fn.lower().endswith('.pdf')):
                continue
            pdf_path = os.path.join(input_dir, fn)
            success, result = PDFExtractor.rename_invoice(pdf_path)
            if success:
                logs.append(f"✓ {fn} → {result}")
                renamed += 1
            else:
                logs.append(f"跳过 {fn}: {result}")

        logs.append(f"重命名完成: {renamed} 个")
        return logs, renamed

    @staticmethod
    def record_sdp_state(downloads_dir=None):
        """Record current sdp_* files in Downloads (for delta detection)."""
        downloads_dir = downloads_dir or Config.DOWNLOADS_DIR
        if os.path.exists(downloads_dir):
            return set(f for f in os.listdir(downloads_dir)
                       if f.startswith('sdp_') and f.lower().endswith('.pdf'))
        return set()

    @staticmethod
    def get_new_sdp_files(before_set, downloads_dir=None):
        """Get newly added sdp_* files since before_set was recorded."""
        downloads_dir = downloads_dir or Config.DOWNLOADS_DIR
        if not os.path.exists(downloads_dir):
            return []
        current = set(f for f in os.listdir(downloads_dir)
                      if f.startswith('sdp_') and f.lower().endswith('.pdf'))
        return [f for f in current if f not in before_set]


# Module-level state for tracking sdp files across API calls
_sdp_before: set = set()


def set_sdp_before(s: set):
    global _sdp_before
    _sdp_before = s


def get_sdp_before() -> set:
    return _sdp_before
